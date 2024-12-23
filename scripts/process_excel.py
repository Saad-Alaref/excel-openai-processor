# scripts/process_excel.py

import os
import time
import logging
import yaml
import pandas as pd
import openai
from pathlib import Path
from typing import Any, Dict, Optional, List
from openpyxl import load_workbook
import json
import operator

# ---------------------------- Logging Configuration ---------------------------- #

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler('../log_file.log'),
        #logging.StreamHandler()  # Added to also output logs to the console
    ]
)

logger = logging.getLogger(__name__)

# ---------------------------- Configuration Loader ---------------------------- #

def load_config(config_path: str) -> Dict[str, Any]:
    try:
        with open(config_path, 'r') as file:
            config = yaml.safe_load(file)
        logger.info(f"Configuration loaded from {config_path}")
        return config
    except Exception as e:
        logger.error(f"Failed to load configuration file: {e}")
        raise

# ---------------------------- OpenAI API Client ---------------------------- #

class OpenAIClient:
    def __init__(self, api_key: str, model: str, system_message: str):
        openai.api_key = api_key  # Set the API key
        self.model = model
        self.system_message = system_message
        logger.info(f"OpenAI client initialized with model '{self.model}' and system message: '{self.system_message}'.")

    def create_completion(
        self,
        prompt: str,
        max_tokens: int,
        temperature: float,
        functions: Optional[List[Dict]] = None,
        function_call: Optional[Dict] = None
    ) -> Optional[Any]:
        try:
            # Prepare the messages
            messages = [
                {"role": "system", "content": self.system_message},
                {"role": "user", "content": prompt}
            ]

            # Create the completion
            completion = openai.chat.completions.create(
                model=self.model,
                messages=messages,
                max_tokens=max_tokens,
                temperature=temperature,
                functions=functions,
                function_call=function_call
            )
            logger.debug("\nCompletion is:")
            logger.debug(completion)
            logger.debug("\n")

            # Handle the response
            message = completion.choices[0].message

            if hasattr(message, 'function_call') and message.function_call:
                # Extract the arguments from the function call
                arguments_str = message.function_call.arguments
                arguments = json.loads(arguments_str)
                return arguments
            else:
                content = message.content.strip() if message.content else ''
                logger.debug(f"OpenAI response content: {content}")
                return content

        except Exception as e:
            logger.error(f"OpenAI API request failed: {e}")
            return None

# ---------------------------- Excel Processor ---------------------------- #

class ExcelProcessor:
    def __init__(self, config: Dict[str, Any], openai_client: OpenAIClient):
        self.config = config
        self.openai = openai_client
        self.input_path = config['excel']['input_path']
        self.sheet_name = config['excel'].get('sheet_name', 'Sheet1')
        self.output_columns = config['columns']['output']
        self.sleep_time = config['processing'].get('sleep_time', 1)
        self.retry_attempts = config['processing'].get('retry_attempts', 3)
        self.retry_delay = config['processing'].get('retry_delay', 5)

        # Load the workbook once to keep it open during processing
        try:
            self.workbook = load_workbook(filename=self.input_path)
            self.sheet = self.workbook[self.sheet_name]
            logger.info(f"Workbook '{self.input_path}' loaded successfully.")
        except Exception as e:
            logger.error(f"Failed to load workbook '{self.input_path}': {e}")
            raise

        # Load filter configuration
        self.filter_enabled = config.get('filter', {}).get('enabled', False)
        self.filter_criteria = config.get('filter', {}).get('criteria', [])

        if self.filter_enabled:
            logger.info("Filtering is enabled. Applying filter criteria.")
        else:
            logger.info("Filtering is disabled. All rows will be processed.")

    def save_workbook(self):
        try:
            self.workbook.save(self.input_path)
            logger.info(f"Workbook '{self.input_path}' saved successfully.")
        except Exception as e:
            logger.error(f"Failed to save workbook '{self.input_path}': {e}")
            raise

    def matches_criteria(self, row_data: pd.Series) -> bool:
        """
        Determine if a row matches all the filter criteria.
        """
        if not self.filter_enabled or not self.filter_criteria:
            return True  # No filtering applied

        # Mapping of operations to actual Python functions
        ops = {
            "equals": operator.eq,
            "contains": lambda a, b: b in a if isinstance(a, str) else False,
            "in": lambda a, b: a in b if a else False,
            "greater_than": operator.gt,
            "less_than": operator.lt,
            # Add more operations as needed
        }

        for criterion in self.filter_criteria:
            column = criterion.get('column')
            operation = criterion.get('operation')
            value = criterion.get('value')

            if column not in row_data:
                logger.warning(f"Filter column '{column}' not found in row data.")
                return False

            cell_value = row_data[column]

            # Handle None or empty cell values
            if pd.isna(cell_value):
                cell_value = None

            # Get the operation function
            op_func = ops.get(operation)
            if not op_func:
                logger.warning(f"Unsupported operation '{operation}' in filter criteria.")
                return False

            # Apply the operation
            try:
                if operation == "in":
                    if not isinstance(value, list):
                        logger.warning(f"Value for 'in' operation must be a list. Got: {value}")
                        return False
                    if not op_func(cell_value, value):
                        return False
                else:
                    if cell_value is None:
                        logger.debug(f"Row has no value for column '{column}'.")
                        return False
                    if not op_func(cell_value, value):
                        return False
            except Exception as e:
                logger.error(f"Error applying filter on column '{column}': {e}")
                return False

            logger.debug(f"Filtering on column '{column}': {cell_value} {operation} {value}")

        return True

    def process_row(self, row_number: int, row_data: pd.Series):
        logger.info(f"Processing row {row_number}: {row_data.to_dict()}")
        for output_column, output_config in self.output_columns.items():
            prompt_template = output_config['prompt']
            max_tokens = output_config.get('max_tokens', 50)
            temperature = output_config.get('temperature', 0.7)
            fetch_all = output_config.get('fetch_all', False)
            schema = output_config.get('schema', None)
            input_columns = output_config.get('input_columns', [])

            # Determine whether to fetch based on 'fetch_all' and cell content
            current_value = row_data.get(output_column, None)
            should_fetch = fetch_all or (pd.isna(current_value) or str(current_value).strip() == '')

            if not should_fetch:
                logger.info(f"Skipping '{output_column}' for row {row_number} as it is already populated and 'fetch_all' is False.")
                continue

            # Prepare the prompt by inserting input column values
            prompt = prompt_template
            for input_col in input_columns:
                prompt += f"\n{input_col}: {row_data[input_col]}"
            logger.debug(f"Generated prompt for '{output_column}': {prompt}")

            # Prepare functions and function_call if schema is provided
            functions = None
            function_call = None
            if schema:
                function_name = schema.get('name', 'auto_generated_function')
                function_schema = schema.get('schema')
                functions = [{
                    "name": function_name,
                    "parameters": function_schema
                }]
                function_call = {"name": function_name}

            # Retry logic
            for attempt in range(1, self.retry_attempts + 1):
                try:
                    result = self.openai.create_completion(
                        prompt=prompt,
                        max_tokens=max_tokens,
                        temperature=temperature,
                        functions=functions,
                        function_call=function_call
                    )

                    logger.info(f"Attempt {attempt}: Received response {result}")

                    if result:
                        if isinstance(result, dict):
                            # When a schema is provided, extract the parameter names
                            if schema:
                                function_schema = schema.get('schema', {})
                                parameter_names = function_schema.get('properties', {}).keys()
                                # Assuming we're interested in the first parameter
                                if parameter_names:
                                    param_name = next(iter(parameter_names))
                                    value = result.get(param_name, 'N/A')
                                    logger.info(f"Extracted '{param_name}': {value}")
                                else:
                                    # If no parameter names are specified, use the entire result
                                    value = result
                                    logger.info(f"No parameters specified in schema, using result: {value}")
                            else:
                                # No schema, use the entire result
                                value = result
                                logger.info(f"No schema provided, using result: {value}")
                        else:
                            # Handle free-form text outputs
                            value = result  # For free-form text outputs
                            logger.info(f"Free-form result: {value}")

                        # Write directly to the cell
                        try:
                            column_letter = self.get_column_letter(output_column)
                            cell_reference = f"{column_letter}{row_number}"
                            self.sheet[cell_reference].value = value
                            logger.info(f" - {output_column} updated in cell {cell_reference}: {value}")
                            break  # Exit retry loop on success
                        except Exception as write_error:
                            logger.error(f"Failed to write to cell {cell_reference}: {write_error}")
                            break  # Break to avoid infinite retries
                    else:
                        logger.warning(f"Attempt {attempt}: Received an empty or invalid result for '{output_column}'.")

                except Exception as e:
                    logger.warning(f"Attempt {attempt} failed for '{output_column}' in row {row_number}. Error: {e}. Retrying in {self.retry_delay} seconds...")
                    time.sleep(self.retry_delay)

            else:
                logger.error(f"All retry attempts failed for '{output_column}' in row {row_number}.")

            # Sleep between API requests to respect rate limits
            time.sleep(self.sleep_time)


    def get_column_letter(self, column_name: str) -> str:
        for idx, cell in enumerate(self.sheet[1], start=1):
            if cell.value == column_name:
                return cell.column_letter
        raise ValueError(f"Column '{column_name}' not found in the Excel sheet.")

    def process_excel(self):
        try:
            total_rows = self.sheet.max_row - 1  # Exclude header row
            current_row = 0
            processed_rows = 0
            print("Processing rows...")
            for idx, row in enumerate(self.sheet.iter_rows(min_row=2, values_only=True), start=2):
                row_data = pd.Series(row, index=[cell.value for cell in self.sheet[1]])

                # Check if the row matches the filter criteria
                if self.matches_criteria(row_data):
                    self.process_row(idx, row_data)
                    processed_rows += 1
                else:
                    logger.debug(f"Row {idx} skipped due to filter criteria.")

                current_row += 1
                percent_complete = (current_row / total_rows) * 100
                print(f"Progress: {current_row}/{total_rows} rows ({percent_complete:.2f}%)", end='\r')
        except Exception as e:
            logger.error(f"Error processing Excel file: {e}")
            raise
        finally:
            # Save the workbook after processing all rows
            self.save_workbook()
            print("\nProcessing complete.")

# ---------------------------- Main Execution ---------------------------- #

def main():

    logger.info("Starting script...")

    # Define paths
    script_dir = Path(__file__).parent
    project_dir = script_dir.parent

    # Load configuration
    config_path = project_dir / 'config' / 'config.yaml'
    config = load_config(str(config_path))

    # Adjust the input_path to be absolute
    excel_input_path = project_dir / config['excel']['input_path']
    config['excel']['input_path'] = str(excel_input_path.resolve())

    # Retrieve OpenAI API key environment variable name from configuration
    api_key_env_var = config['openai']['api_key_env_var']
    api_key = os.getenv(api_key_env_var)
    if not api_key:
        logger.error(f"OpenAI API key not found. Please set the '{api_key_env_var}' environment variable.")
        return

    # Retrieve OpenAI model and system message from configuration
    model = config['openai'].get('model', 'gpt-4o-mini')
    system_message = config['openai'].get('system_message', "You are a helpful assistant that adheres to user requests.")

    # Initialize OpenAI client
    openai_client = OpenAIClient(api_key=api_key, model=model, system_message=system_message)

    # Initialize Excel processor
    processor = ExcelProcessor(config=config, openai_client=openai_client)

    # Process Excel file
    processor.process_excel()

    logger.info("Excel processing completed successfully.\nExiting...")

if __name__ == "__main__":
    main()
